"""
Excel Parser — reads .xlsx upload and returns rows as list-of-dicts.

openpyxl returns typed cell values (datetime, int, float, bool) when using
read_only=True + values_only=True.  _cell_to_str() normalises all of them
to clean strings before they enter the SQL generator.
"""
import io
import re
import datetime as dt
from typing import Tuple

import openpyxl

_MULTI_SPACE = re.compile(r" {2,}")


def _cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "1" if val else "0"
    if isinstance(val, dt.datetime):
        if val.hour == 0 and val.minute == 0 and val.second == 0 and val.microsecond == 0:
            return val.strftime("%Y-%m-%d")
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, dt.date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        return str(int(val)) if val.is_integer() else str(val)
    if isinstance(val, int):
        return str(val)
    return _MULTI_SPACE.sub(" ", str(val).strip())


def _parse_headers(row_cells: tuple) -> list[str]:
    headers: list[str] = []
    for idx, h in enumerate(row_cells):
        cleaned = _cell_to_str(h).strip()
        headers.append(cleaned if cleaned else f"Column_{idx + 1}")
    return headers


def read_header_row(file_content: bytes) -> list[str]:
    """Read ONLY the first row without loading all data rows."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open workbook: {exc}") from exc

    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Workbook appears to be empty.")

    try:
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = _parse_headers(row)
            if not headers:
                raise ValueError("No header row found in the uploaded file.")
            return headers
    finally:
        wb.close()

    raise ValueError("No header row found in the uploaded file.")


def parse_upload_excel(file_content: bytes, filename: str) -> Tuple[list[dict], list[str]]:
    """Parse an uploaded Excel file and return (rows, columns)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open workbook: {exc}") from exc

    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Workbook appears to be empty or has no data rows.")

        columns: list[str] = []
        rows: list[dict] = []

        for row_idx, row_cells in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                columns = _parse_headers(row_cells)
                continue

            if all(c is None or _cell_to_str(c) == "" for c in row_cells):
                continue

            n = min(len(columns), len(row_cells))
            rows.append({columns[i]: _cell_to_str(row_cells[i]) for i in range(n)})

        if not columns:
            raise ValueError("No header row found in the uploaded file.")
        if not rows:
            raise ValueError("No data rows found in the uploaded file.")

        return rows, columns

    finally:
        wb.close()
