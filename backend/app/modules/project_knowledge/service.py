from __future__ import annotations

import csv
import hashlib
import logging
import re
import zipfile
from collections import Counter
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import fitz
from openpyxl import load_workbook

from app.config.local_db import connection, json_text, row_to_dict
from app.config.settings import settings
from app.modules.instruments.service import ensure_project
from app.modules.llm.service import _is_available, generate

logger = logging.getLogger(__name__)

PROJECT_KNOWLEDGE_MODEL = "xyra-project-context"
SUPPORTED_EXTENSIONS = {".pdf", ".txt", ".md", ".csv", ".xlsx", ".xlsm", ".docx"}
STOPWORDS = {
    "a", "an", "and", "are", "as", "at", "be", "by", "can", "do", "for", "from", "has", "in", "is", "it",
    "of", "on", "or", "shall", "should", "that", "the", "this", "to", "with", "what", "when", "where",
}

QUERY_EXPANSIONS = {
    "pid": ["p&id", "piping instrumentation", "piping and instrumentation", "drawing"],
    "p&id": ["pid", "piping instrumentation", "piping and instrumentation", "drawing"],
    "instrumentation": ["instrument", "instrumentation", "isa"],
    "instrument": ["instrumentation", "instrument"],
    "symbol": ["symbols", "legend", "symbology"],
    "symbols": ["symbol", "legend", "symbology"],
    "standard": ["standards", "specification", "specifications", "spec"],
    "standards": ["standard", "specification", "specifications", "spec"],
    "shutdown": ["sdv", "esd", "trip", "interlock", "shut down"],
    "valve": ["valves", "valve"],
    "valves": ["valve", "valves"],
    "mto": ["material take-off", "material take off", "take-off", "takeoff"],
}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _s(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _safe_int(value: Any, default: int = 1) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _hash_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _tokens(text: str) -> list[str]:
    text = (text or "").lower().replace("p&id", "pid")
    raw = [t for t in re.findall(r"[A-Za-z0-9][A-Za-z0-9_./#-]{1,}", text) if t not in STOPWORDS]
    tokens = []
    for token in raw:
        tokens.append(token)
        if len(token) > 4 and token.endswith("s"):
            tokens.append(token[:-1])
    return tokens


def _expanded_terms(query: str) -> list[str]:
    ordered: list[str] = []
    for term in _tokens(query):
        for candidate in [term, *QUERY_EXPANSIONS.get(term, [])]:
            candidate = candidate.lower().strip()
            if candidate and candidate not in ordered and candidate not in STOPWORDS:
                ordered.append(candidate)
    return ordered


def _term_weights(query: str, terms: list[str]) -> dict[str, int]:
    original_terms = set(_tokens(query))
    weights: dict[str, int] = {}
    for term in terms:
        weight = 2 if term in original_terms else 1
        if re.search(r"\d", term):
            weight += 2
        if term in {"iec", "isa", "api", "asme", "ansi", "61511", "standard", "standards", "specification"}:
            weight += 1
        if len(term) >= 8 and " " not in term:
            weight += 1
        weights[term] = weight
    return weights


def _clean_text(text: str) -> str:
    text = unescape(text or "")
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _chunk_page(text: str, *, max_chars: int = 1400, overlap: int = 180) -> list[str]:
    text = _clean_text(text)
    if not text:
        return []
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    chunks: list[str] = []
    current = ""
    for para in paragraphs or [text]:
        if len(current) + len(para) + 2 <= max_chars:
            current = f"{current}\n\n{para}".strip()
            continue
        if current:
            chunks.append(current)
        if len(para) <= max_chars:
            current = para
            continue
        start = 0
        while start < len(para):
            end = min(len(para), start + max_chars)
            chunks.append(para[start:end].strip())
            if end == len(para):
                break
            start = max(0, end - overlap)
        current = ""
    if current:
        chunks.append(current)
    return chunks


def _extract_pdf(path: Path) -> list[dict]:
    pages = []
    doc = fitz.open(path)
    try:
        for index, page in enumerate(doc, start=1):
            text = _clean_text(page.get_text("text"))
            if text:
                pages.append({"page": index, "section": "", "text": text})
    finally:
        doc.close()
    return pages


def _extract_text(path: Path) -> list[dict]:
    return [{"page": None, "section": "", "text": _clean_text(path.read_text(encoding="utf-8", errors="ignore"))}]


def _extract_csv(path: Path) -> list[dict]:
    rows = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as fh:
        reader = csv.reader(fh)
        for idx, row in enumerate(reader, start=1):
            if idx > 5000:
                break
            rows.append(" | ".join(_s(cell) for cell in row if _s(cell)))
    return [{"page": None, "section": path.name, "text": _clean_text("\n".join(rows))}]


def _extract_excel(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    pages = []
    try:
        for ws in wb.worksheets:
            lines = []
            for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if ridx > 1500:
                    lines.append("[truncated after 1500 rows]")
                    break
                cells = [_s(cell) for cell in row if _s(cell)]
                if cells:
                    lines.append(" | ".join(cells))
            text = _clean_text("\n".join(lines))
            if text:
                pages.append({"page": None, "section": ws.title, "text": text})
    finally:
        wb.close()
    return pages


def _extract_docx(path: Path) -> list[dict]:
    texts: list[str] = []
    with zipfile.ZipFile(path) as zf:
        for name in ("word/document.xml",):
            if name not in zf.namelist():
                continue
            root = ElementTree.fromstring(zf.read(name))
            for node in root.iter():
                if node.tag.endswith("}t") and node.text:
                    texts.append(node.text)
                elif node.tag.endswith("}p"):
                    texts.append("\n")
    return [{"page": None, "section": "", "text": _clean_text(" ".join(texts).replace(" \n ", "\n"))}]


def _extract_document(path: Path) -> list[dict]:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return _extract_pdf(path)
    if ext in {".txt", ".md"}:
        return _extract_text(path)
    if ext == ".csv":
        return _extract_csv(path)
    if ext in {".xlsx", ".xlsm"}:
        return _extract_excel(path)
    if ext == ".docx":
        return _extract_docx(path)
    return []


def _classify_document(path: Path, text_sample: str) -> str:
    hay = f"{path.name} {text_sample[:3000]}".lower()
    if "scope" in hay or "sow" in hay or "work scope" in hay:
        return "SOW"
    if "standard" in hay or "specification" in hay or "spec" in hay:
        return "Specification"
    if "p&id" in hay or "piping and instrumentation" in hay:
        return "P&ID / drawing"
    if "datasheet" in hay or "data sheet" in hay:
        return "Datasheet"
    if "line list" in hay:
        return "Line list"
    if "instrument index" in hay:
        return "Instrument index"
    return "Project document"


def discover_files(folder_path: str) -> list[Path]:
    root = Path(folder_path).expanduser()
    if not root.exists() or not root.is_dir():
        raise ValueError(f"Folder does not exist: {folder_path}")
    files = [
        path
        for path in root.rglob("*")
        if path.is_file()
        and path.suffix.lower() in SUPPORTED_EXTENSIONS
        and not any(part.startswith(".") for part in path.parts)
        and "__pycache__" not in path.parts
    ]
    return sorted(files, key=lambda p: str(p).lower())


def index_folder(project_id: str, folder_path: str, *, force: bool = False, limit: int = 300) -> dict:
    project_id = _s(project_id) or settings.XYRA_DEFAULT_PROJECT_ID or "default"
    ensure_project(project_id)
    files = discover_files(folder_path)[: max(1, min(limit, 2000))]
    stats = {"project_id": project_id, "folder_path": str(Path(folder_path).expanduser()), "files_seen": len(files), "indexed": 0, "skipped": 0, "failed": 0, "chunks": 0, "errors": []}

    with connection() as conn:
        for path in files:
            try:
                file_hash = _hash_file(path)
                existing = conn.execute(
                    "SELECT id, file_hash FROM project_knowledge_documents WHERE project_id=? AND file_path=?",
                    (project_id, str(path)),
                ).fetchone()
                if existing and existing["file_hash"] == file_hash and not force:
                    stats["skipped"] += 1
                    continue

                extracted = _extract_document(path)
                full_sample = "\n".join(page["text"][:1500] for page in extracted[:3])
                doc_type = _classify_document(path, full_sample)
                now = _now()
                if existing:
                    doc_id = existing["id"]
                    conn.execute(
                        """
                        UPDATE project_knowledge_documents
                        SET file_name=?, file_hash=?, file_size_bytes=?, extension=?, document_type=?,
                            status='indexed', error_message='', indexed_at=?, updated_at=?
                        WHERE id=?
                        """,
                        (path.name, file_hash, path.stat().st_size, path.suffix.lower(), doc_type, now, now, doc_id),
                    )
                    conn.execute("DELETE FROM project_knowledge_chunks WHERE document_id=?", (doc_id,))
                else:
                    cur = conn.execute(
                        """
                        INSERT INTO project_knowledge_documents (
                            project_id, folder_path, file_path, file_name, file_hash, file_size_bytes,
                            extension, document_type, status, indexed_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'indexed', ?, ?)
                        """,
                        (project_id, str(Path(folder_path).expanduser()), str(path), path.name, file_hash, path.stat().st_size, path.suffix.lower(), doc_type, now, now),
                    )
                    doc_id = conn.execute("SELECT id FROM project_knowledge_documents WHERE rowid=?", (cur.lastrowid,)).fetchone()["id"]

                chunk_count = 0
                for page in extracted:
                    chunks = _chunk_page(page["text"])
                    for local_index, chunk in enumerate(chunks):
                        token_counts = Counter(_tokens(chunk))
                        conn.execute(
                            """
                            INSERT INTO project_knowledge_chunks (
                                project_id, document_id, chunk_index, page_number, section_title,
                                content, token_count, keywords, created_at
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                project_id,
                                doc_id,
                                chunk_count + local_index,
                                page.get("page"),
                                page.get("section") or "",
                                chunk,
                                len(chunk.split()),
                                json_text([word for word, _ in token_counts.most_common(24)], []),
                                now,
                            ),
                        )
                    chunk_count += len(chunks)
                conn.execute(
                    "UPDATE project_knowledge_documents SET chunk_count=?, status=? WHERE id=?",
                    (chunk_count, "indexed" if chunk_count else "empty", doc_id),
                )
                stats["indexed"] += 1
                stats["chunks"] += chunk_count
            except Exception as exc:
                logger.exception("Project knowledge indexing failed for %s", path)
                stats["failed"] += 1
                stats["errors"].append({"file": str(path), "error": str(exc)[:300]})
    return stats


def list_documents(project_id: str) -> dict:
    project_id = _s(project_id) or settings.XYRA_DEFAULT_PROJECT_ID or "default"
    with connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM project_knowledge_documents
            WHERE project_id=?
            ORDER BY indexed_at DESC, file_name
            """,
            (project_id,),
        ).fetchall()
    docs = [row_to_dict(row) or {} for row in rows]
    return {
        "project_id": project_id,
        "documents": docs,
        "total": len(docs),
        "chunks": sum(int(doc.get("chunk_count") or 0) for doc in docs),
    }


def render_page_image(project_id: str, document_id: str, *, page_number: int = 1, zoom: float = 0.75) -> bytes:
    project_id = _s(project_id) or settings.XYRA_DEFAULT_PROJECT_ID or "default"
    document_id = _s(document_id)
    if not document_id:
        raise ValueError("document_id is required")
    with connection() as conn:
        row = conn.execute(
            """
            SELECT file_path, extension
            FROM project_knowledge_documents
            WHERE id=? AND project_id=?
            """,
            (document_id, project_id),
        ).fetchone()
    if not row:
        raise ValueError("Document not found")
    path = Path(row["file_path"])
    if row["extension"] != ".pdf" or not path.exists():
        raise ValueError("Preview is available only for indexed PDF documents")
    doc = fitz.open(path)
    try:
        page_index = max(0, min(_safe_int(page_number, 1) - 1, len(doc) - 1))
        matrix = fitz.Matrix(max(0.35, min(float(zoom), 2.0)), max(0.35, min(float(zoom), 2.0)))
        pix = doc[page_index].get_pixmap(matrix=matrix, alpha=False)
        return pix.tobytes("png")
    finally:
        doc.close()


def list_saved_evidence(project_id: str) -> dict:
    project_id = _s(project_id) or settings.XYRA_DEFAULT_PROJECT_ID or "default"
    with connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM project_knowledge_saved_evidence
            WHERE project_id=?
            ORDER BY created_at DESC
            LIMIT 200
            """,
            (project_id,),
        ).fetchall()
    items = [row_to_dict(row) or {} for row in rows]
    return {"project_id": project_id, "items": items, "total": len(items)}


def save_evidence(project_id: str, citation: dict, *, question: str = "", note: str = "") -> dict:
    project_id = _s(project_id) or settings.XYRA_DEFAULT_PROJECT_ID or "default"
    ensure_project(project_id)
    citation = citation or {}
    chunk_id = _s(citation.get("chunk_id"))
    document_id = _s(citation.get("document_id"))
    if not chunk_id and not document_id:
        raise ValueError("citation must include chunk_id or document_id")
    with connection() as conn:
        if chunk_id:
            row = conn.execute(
                "SELECT id FROM project_knowledge_saved_evidence WHERE project_id=? AND chunk_id=?",
                (project_id, chunk_id),
            ).fetchone()
        else:
            row = conn.execute(
                "SELECT id FROM project_knowledge_saved_evidence WHERE project_id=? AND document_id=? AND chunk_id IS NULL",
                (project_id, document_id),
            ).fetchone()
        if row:
            conn.execute(
                """
                UPDATE project_knowledge_saved_evidence
                SET question=?, note=?, citation_snapshot=?
                WHERE id=?
                """,
                (_s(question), _s(note), json_text(citation, {}), row["id"]),
            )
            saved_id = row["id"]
        else:
            cur = conn.execute(
                """
                INSERT INTO project_knowledge_saved_evidence (
                    project_id, chunk_id, document_id, question, note, citation_snapshot
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (project_id, chunk_id or None, document_id or None, _s(question), _s(note), json_text(citation, {})),
            )
            saved_id = conn.execute("SELECT id FROM project_knowledge_saved_evidence WHERE rowid=?", (cur.lastrowid,)).fetchone()["id"]
        saved = conn.execute("SELECT * FROM project_knowledge_saved_evidence WHERE id=?", (saved_id,)).fetchone()
    return row_to_dict(saved) or {}


def delete_saved_evidence(project_id: str, saved_id: str) -> dict:
    project_id = _s(project_id) or settings.XYRA_DEFAULT_PROJECT_ID or "default"
    saved_id = _s(saved_id)
    with connection() as conn:
        cur = conn.execute(
            "DELETE FROM project_knowledge_saved_evidence WHERE project_id=? AND id=?",
            (project_id, saved_id),
        )
    return {"project_id": project_id, "id": saved_id, "deleted": cur.rowcount > 0}


def search(project_id: str, query: str, *, limit: int = 10, document_type: str = "") -> dict:
    project_id = _s(project_id) or settings.XYRA_DEFAULT_PROJECT_ID or "default"
    query = _s(query)
    terms = _expanded_terms(query)
    if not terms:
        return {"project_id": project_id, "query": query, "results": []}
    weights = _term_weights(query, terms)
    query_phrases = [phrase.lower().strip() for phrase in re.split(r"\s{2,}|\?|,", query) if len(phrase.strip()) > 8]
    limit = max(1, min(limit, 50))

    where = ["c.project_id=?"]
    params: list[Any] = [project_id]
    if document_type:
        where.append("d.document_type=?")
        params.append(document_type)
    like_terms = terms[:12]
    where.append("(" + " OR ".join("lower(c.content) LIKE ?" for _ in like_terms) + ")")
    params.extend([f"%{term}%" for term in like_terms])

    with connection() as conn:
        rows = conn.execute(
            f"""
            SELECT c.*, d.file_name, d.file_path, d.document_type, d.extension
            FROM project_knowledge_chunks c
            JOIN project_knowledge_documents d ON d.id=c.document_id
            WHERE {' AND '.join(where)}
            ORDER BY d.file_name, c.page_number, c.chunk_index
            LIMIT 400
            """,
            tuple(params),
        ).fetchall()

    scored = []
    for row in rows:
        data = row_to_dict(row) or {}
        content = _s(data.get("content"))
        content_l = content.lower()
        keywords = set(data.get("keywords") or [])
        file_l = _s(data.get("file_name")).lower()
        doc_type_l = _s(data.get("document_type")).lower()
        matched_terms = {
            term
            for term in terms
            if term in content_l or term in file_l or term in doc_type_l or term in keywords
        }
        weighted_hits = sum(min(content_l.count(term), 6) * weights.get(term, 1) for term in terms)
        coverage = sum(weights.get(term, 1) for term in matched_terms)
        original_matched = {term for term in _tokens(query) if term in matched_terms}
        score = weighted_hits
        score += coverage * 7
        score += max(0, len(matched_terms) - 1) * 3
        score += len(original_matched) * 10
        exact_bonus = 18 if query.lower() in content_l else 0
        phrase_bonus = sum(10 for phrase in query_phrases if phrase in content_l)
        title_bonus = sum(6 for term in matched_terms if term in file_l)
        doc_type_bonus = sum(5 for term in matched_terms if term in doc_type_l)
        early_bonus = max(0, 5 - int((data.get("page_number") or 1) / 20))
        score = score + exact_bonus + phrase_bonus + title_bonus
        score = score + doc_type_bonus + early_bonus
        if score <= 0:
            continue
        data["score"] = score
        data["matched_terms"] = sorted(matched_terms)
        data["excerpt"] = _excerpt(content, terms)
        scored.append(data)
    scored.sort(key=lambda item: (-item["score"], item.get("file_name") or "", item.get("page_number") or 0))
    return {"project_id": project_id, "query": query, "results": [_citation(row) for row in scored[:limit]]}


def _excerpt(content: str, terms: list[str], radius: int = 260) -> str:
    lower = content.lower()
    positions = [lower.find(term) for term in terms if lower.find(term) >= 0]
    if not positions:
        return content[: radius * 2].strip()
    center = min(positions)
    start = max(0, center - radius)
    end = min(len(content), center + radius)
    return content[start:end].strip()


def _citation(row: dict) -> dict:
    return {
        "chunk_id": row.get("id"),
        "document_id": row.get("document_id"),
        "file_name": row.get("file_name"),
        "file_path": row.get("file_path"),
        "document_type": row.get("document_type"),
        "page_number": row.get("page_number"),
        "section_title": row.get("section_title"),
        "score": row.get("score"),
        "matched_terms": row.get("matched_terms") or [],
        "excerpt": row.get("excerpt") or _s(row.get("content"))[:600],
    }


def _rule_answer(question: str, citations: list[dict]) -> str:
    if not citations:
        return "I could not find matching project knowledge yet. Index the project folder or broaden the question."
    high_value_terms = []
    for item in citations:
        for term in item.get("matched_terms") or []:
            if term not in high_value_terms and len(term) > 2:
                high_value_terms.append(term)
    lines = [
        "Based on the indexed project documents:",
        "",
        "What I found",
    ]
    for idx, item in enumerate(citations[:5], start=1):
        page = f" p.{item['page_number']}" if item.get("page_number") else ""
        excerpt = item["excerpt"][:300].replace(chr(10), " ")
        lines.append(f"{idx}. {item['file_name']}{page}: {excerpt}")
    if high_value_terms:
        lines.extend(["", f"Matched engineering terms: {', '.join(high_value_terms[:12])}."])
    lines.extend([
        "",
        "Action for engineer: open the cited pages before issuing a deliverable, especially if the answer affects scope, design basis, safety, or client compliance.",
    ])
    return "\n".join(lines)


def _suggest_followups(question: str, citations: list[dict]) -> list[str]:
    terms: list[str] = []
    for item in citations:
        for term in item.get("matched_terms") or []:
            if term not in terms and len(term) > 2:
                terms.append(term)
    base = [
        "Which cited pages contain mandatory client requirements?",
        "What information is missing before issuing the deliverable?",
        "Summarize only the requirements and ignore generic background text.",
    ]
    if any(term in {"valve", "valves", "mto", "piping"} for term in terms):
        base.insert(0, "Create a piping/MTO checklist from the cited evidence.")
    if any(term in {"instrument", "instrumentation", "isa", "symbol", "symbols"} for term in terms):
        base.insert(0, "Extract instrumentation rules, symbols, and tagging guidance from the citations.")
    if "scope" in _tokens(question):
        base.insert(0, "Separate project scope, exclusions, and assumptions from the cited documents.")
    return base[:5]


async def chat(project_id: str, question: str, *, limit: int = 8, use_model: bool = True) -> dict:
    hits = search(project_id, question, limit=limit)
    citations = hits["results"]
    answer = _rule_answer(question, citations)
    model_status = {"model": PROJECT_KNOWLEDGE_MODEL, "status": "not_used"}
    if use_model and citations and _is_available(PROJECT_KNOWLEDGE_MODEL):
        context = "\n\n".join(
            f"[{idx}] {c['file_name']} page {c.get('page_number') or '-'}\n{c['excerpt']}"
            for idx, c in enumerate(citations, start=1)
        )
        prompt = f"""
You are XYRA Project Knowledge, an offline EPC project assistant.
Answer only from the provided project document excerpts.
If the excerpts do not contain the answer, say what is missing.
Write a practical engineering answer: direct answer first, then requirements/evidence, then gaps or checks.
Return JSON with keys: answer, confidence, follow_up_questions.

Question:
{question}

Project excerpts:
{context}
"""
        data = generate(prompt, model=PROJECT_KNOWLEDGE_MODEL, timeout=90, num_predict=900)
        if data and _s(data.get("answer")):
            answer = _s(data.get("answer"))
            model_status = {
                "model": PROJECT_KNOWLEDGE_MODEL,
                "status": "answered",
                "confidence": data.get("confidence"),
                "follow_up_questions": data.get("follow_up_questions") if isinstance(data.get("follow_up_questions"), list) else [],
            }
        else:
            model_status = {"model": PROJECT_KNOWLEDGE_MODEL, "status": "fallback_rules"}
    elif use_model:
        model_status = {"model": PROJECT_KNOWLEDGE_MODEL, "status": "unavailable_or_no_citations"}
    result = {
        "project_id": project_id,
        "question": question,
        "answer": answer,
        "citations": citations,
        "model_status": model_status,
    }
    followups = model_status.get("follow_up_questions") if isinstance(model_status.get("follow_up_questions"), list) else []
    result["follow_up_questions"] = followups or _suggest_followups(question, citations)
    with connection() as conn:
        conn.execute(
            """
            INSERT INTO project_knowledge_chat (project_id, question, answer, citations, model, status)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                project_id,
                question,
                answer,
                json_text(citations, []),
                model_status.get("model"),
                model_status.get("status"),
            ),
        )
    return result
