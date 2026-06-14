import zipfile
from copy import deepcopy

from openpyxl import load_workbook

from app.modules.piping_mto.excel_writer import write_mto_package


def _payload():
    return {
        "project": {
            "project_name": "Install Flowmeters on Flares",
            "project_no": "50213-MTO-00-LA-2615",
            "client_name": "ADNOC",
            "contractor_name": "Bilfinger Tebodin",
            "location": "Abu Dhabi",
        },
        "threshold": 0.7,
        "sessions": [
            {
                "id": "ball",
                "label": "Ball valve",
                "count": 3,
                "metadata": {
                    "categoryCode": "H",
                    "categoryName": "BALL VALVE",
                    "unit": "-",
                    "itemType": "Ball valve",
                    "pipingClass": "AC1A1B-FA",
                    "sizeInch": "2",
                    "rating": "150#",
                    "valveBore": "FB",
                    "endConnection": "FLANGED",
                    "materialDescription": "BALL VALVE, FULL BORE, FLOATING BALL, CL150",
                    "dataSheetDocumentNo": "50213-DS-00-LA-8004",
                    "dataSheetReferenceNo": "BV-010",
                    "remarks": "",
                },
                "fileResults": [
                    {
                        "fileName": "PID-001.pdf",
                        "count": 3,
                        "pageCounts": [{"page": 1, "count": 2}, {"page": 2, "count": 1}],
                        "imageWidth": 1000,
                        "imageHeight": 800,
                        "matches": [
                            {"page": 1, "x1": 1, "y1": 2, "x2": 10, "y2": 20, "score": 0.91},
                            {"page": 1, "x1": 11, "y1": 12, "x2": 20, "y2": 30, "score": 0.89},
                            {"page": 2, "x1": 21, "y1": 22, "x2": 30, "y2": 40, "score": 0.81},
                        ],
                    }
                ],
            }
        ],
    }


def test_write_mto_package_creates_expected_zip_and_workbook(tmp_path):
    zip_path = write_mto_package(tmp_path, _payload(), "mto_test")

    assert zip_path.exists()
    with zipfile.ZipFile(zip_path) as zf:
        names = set(zf.namelist())
    assert "Piping Material Take-Off.xlsx" in names
    assert "Detection Register.xlsx" in names
    assert "QA Checks.xlsx" in names
    assert "MTO Engineering Audit.xlsx" in names
    assert "mto_run.json" in names

    wb = load_workbook(tmp_path / "Piping Material Take-Off.xlsx", data_only=True)
    ws = wb["Valves MTO"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "PIPING MATERIAL TAKE-OFF (VALVES MTO)" in values
    assert "H. BALL VALVE" in values
    assert "Ball valve" in values
    assert 3 in values


def test_detection_register_preserves_page_numbers(tmp_path):
    write_mto_package(tmp_path, _payload(), "mto_test")
    wb = load_workbook(tmp_path / "Detection Register.xlsx", data_only=True)
    ws = wb["Detection Register"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    page_col = headers.index("Page") + 1
    pages = [ws.cell(row=i, column=page_col).value for i in range(2, 5)]
    assert pages == [1, 1, 2]


def test_deselected_mto_detection_is_registered_but_not_counted(tmp_path):
    payload = deepcopy(_payload())
    payload["sessions"][0]["fileResults"][0]["matches"][1]["accepted"] = False

    write_mto_package(tmp_path, payload, "mto_test")

    register_wb = load_workbook(tmp_path / "Detection Register.xlsx", data_only=True)
    register_ws = register_wb["Detection Register"]
    headers = [register_ws.cell(row=1, column=i).value for i in range(1, register_ws.max_column + 1)]
    selected_col = headers.index("Selected") + 1
    assert [register_ws.cell(row=i, column=selected_col).value for i in range(2, 5)] == ["Yes", "No", "Yes"]

    mto_wb = load_workbook(tmp_path / "Piping Material Take-Off.xlsx", data_only=True)
    mto_ws = mto_wb["Valves MTO"]
    values = [cell.value for row in mto_ws.iter_rows() for cell in row]
    assert 2 in values
    assert 3 not in values

    audit_wb = load_workbook(tmp_path / "MTO Engineering Audit.xlsx", data_only=True)
    audit_ws = audit_wb["Detection Audit"]
    audit_headers = [audit_ws.cell(row=1, column=i).value for i in range(1, audit_ws.max_column + 1)]
    grade_col = audit_headers.index("Grade") + 1
    assert [audit_ws.cell(row=i, column=grade_col).value for i in range(2, 5)] == ["Review", "Excluded", "Review"]


def test_mto_material_description_falls_back_to_item_type_and_size(tmp_path):
    payload = deepcopy(_payload())
    metadata = payload["sessions"][0]["metadata"]
    metadata["materialDescription"] = ""
    metadata["sizeInch"] = ""
    payload["sessions"][0]["fileResults"][0]["matches"][0]["sizeInch"] = "0.75"
    payload["sessions"][0]["fileResults"][0]["matches"] = payload["sessions"][0]["fileResults"][0]["matches"][:1]

    write_mto_package(tmp_path, payload, "mto_test")

    wb = load_workbook(tmp_path / "Piping Material Take-Off.xlsx", data_only=True)
    ws = wb["Valves MTO"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "BALL VALVE, 0.75 IN" in values
