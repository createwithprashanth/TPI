from __future__ import annotations

import json
from pathlib import Path
from typing import Any


def latest_mto_export(export_dir: Path) -> Path | None:
    if not export_dir.exists():
        return None
    candidates = sorted(
        [path for path in export_dir.iterdir() if path.is_dir()],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def read_mto_rows(export_dir: Path, limit: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    run_dir = latest_mto_export(export_dir)
    if not run_dir:
        return [], {"source": "none", "message": "No MTO export folder found"}

    run_json = run_dir / "mto_run.json"
    if run_json.exists():
        try:
            payload = json.loads(run_json.read_text(encoding="utf-8"))
            rows = _rows_from_mto_json(payload, limit)
            if rows:
                return rows, {"source": str(run_json), "run_dir": str(run_dir)}
        except Exception as exc:
            json_error = str(exc)
        else:
            json_error = "summary-only JSON; fell back to workbook"

    workbook = run_dir / "Piping Material Take-Off.xlsx"
    if workbook.exists():
        try:
            rows = _rows_from_workbook(workbook, limit)
            return rows, {"source": str(workbook), "run_dir": str(run_dir), "json_note": locals().get("json_error", "")}
        except Exception as exc:
            return [], {"source": str(workbook), "error": str(exc)}

    return [], {"source": str(run_dir), "message": "No readable MTO JSON/workbook found"}


def _rows_from_mto_json(payload: dict[str, Any], limit: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    sessions = payload.get("sessions", []) if isinstance(payload, dict) else []
    for session in sessions:
        metadata = session.get("metadata", {}) if isinstance(session, dict) else {}
        for file_result in session.get("fileResults", []) or []:
            for match in file_result.get("matches", []) or []:
                rows.append(
                    {
                        "component_id": f"{session.get('id','')}-{len(rows)+1}",
                        "file": file_result.get("fileName", ""),
                        "page": match.get("page", 1),
                        "item_type": metadata.get("itemType") or session.get("label", ""),
                        "piping_class": metadata.get("pipingClass", ""),
                        "size_inch": match.get("aiNormalizedSizeInch") or match.get("sizeInch") or metadata.get("sizeInch", ""),
                        "rating": metadata.get("rating", ""),
                        "valve_bore": metadata.get("valveBore", ""),
                        "end_connection": metadata.get("endConnection", ""),
                        "quantity": 1,
                        "score": match.get("score", 0),
                        "ai_decision": match.get("aiDecision", ""),
                        "ai_reason": match.get("aiReason", ""),
                        "line_number": match.get("aiLineNumber", ""),
                        "material_description": metadata.get("materialDescription") or match.get("aiMaterialDescriptionHint", ""),
                    }
                )
                if len(rows) >= limit:
                    return rows
    return rows


def _rows_from_workbook(path: Path, limit: int) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    headers: list[str] = []
    for row in ws.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if not any(values):
            continue
        lower = [value.lower() for value in values]
        normalized = [_normalize_header(value) for value in values]
        if "item_type" in normalized and "total_quantity" in normalized:
            headers = normalized
            continue
        if not headers or values[0].lower().startswith(("h. ball", "i.", "j.", "k.")):
            continue
        item = {headers[i] if i < len(headers) else f"col_{i}": values[i] for i in range(len(values))}
        if item.get("item_type") or item.get("material_description"):
            rows.append(
                {
                    "component_id": item.get("sl_no", f"mto-{len(rows)+1}"),
                    "unit": item.get("unit", ""),
                    "item_type": item.get("item_type", ""),
                    "piping_class": item.get("piping_class", ""),
                    "size_inch": item.get("size_inch", ""),
                    "rating": item.get("rating", ""),
                    "valve_bore": item.get("valve_bore", ""),
                    "end_connection": item.get("end_connection", ""),
                    "material_description": item.get("material_description", ""),
                    "datasheet_document_no": item.get("data_sheet_document_nos", ""),
                    "datasheet_reference_no": item.get("data_sheet_reference_nos", ""),
                    "quantity": item.get("total_quantity", ""),
                    "remarks": item.get("remarks", ""),
                }
            )
        if len(rows) >= limit:
            break
    return rows


def _normalize_header(value: str) -> str:
    text = value.strip().lower().replace("\n", " ")
    replacements = {
        "sl. no.": "sl_no",
        "unit": "unit",
        "item type": "item_type",
        "piping class": "piping_class",
        "size (inch)": "size_inch",
        "rating": "rating",
        "valve bore": "valve_bore",
        "end connection": "end_connection",
        "material description": "material_description",
        "data sheet document nos.": "data_sheet_document_nos",
        "data sheet reference nos.": "data_sheet_reference_nos",
        "total quantity": "total_quantity",
        "remarks": "remarks",
    }
    return replacements.get(text, text.replace(" ", "_").replace(".", ""))
